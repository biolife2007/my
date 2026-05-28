import csv
import json
import re
import shutil
import sys
import time
import zipfile
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urljoin, urlparse, urlunparse
import xml.etree.ElementTree as ET

import openpyxl
import requests
from bs4 import BeautifulSoup


BASE = "https://hmel-master.com"
PRICE_FILE = Path(r"C:\Work\ПОСТАЧАЛЬНИКИ\Магнум Стіл\Парсинг з файлу\Прайс Партнер MagnumStill.xlsx")
OUT_DIR = Path.cwd() / "tmp_magnum_match"
CATALOG_JSON = OUT_DIR / "hmel_master_catalog.json"
MATCH_CSV = OUT_DIR / "match_report.csv"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


@dataclass(frozen=True)
class Product:
    title: str
    sku: str
    url: str


def norm(text):
    if text is None:
        return ""
    text = str(text).lower().replace("ё", "е").replace("’", "'").replace("“", '"').replace("”", '"')
    text = text.replace("під тен", "з клампом під тен")
    text = text.replace("без тену", "без тена").replace("без тэна", "без тена")
    text = text.replace("магнум", "magnum").replace("фірмовий", "фирмовий")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[\(\)\[\],.;:!]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def token_set(text):
    return set(re.findall(r"[a-zа-яіїєґ0-9]+|[234]\"", norm(text), re.I))


def volumes(text):
    return set(re.findall(r"(\d+(?:[,.]\d+)?)\s*л\b", norm(text)))


def inches(text):
    return set(re.findall(r'([123468])\s*"', norm(text)))


def score(query, product):
    qn = norm(query)
    pn = norm(product.title)
    if not qn or not pn:
        return 0.0
    if qn == pn:
        return 1.0
    q_tokens = token_set(query)
    p_tokens = token_set(product.title)
    overlap = len(q_tokens & p_tokens) / max(1, len(q_tokens))
    seq = SequenceMatcher(None, qn, pn).ratio()
    contains = 1.0 if qn in pn or pn in qn else 0.0
    result = max(seq, 0.70 * overlap + 0.30 * seq, 0.92 * contains + 0.08 * seq)
    q_volumes = volumes(query)
    p_volumes = volumes(product.title)
    q_inches = inches(query)
    p_inches = inches(product.title)
    if q_volumes and p_volumes and not (q_volumes & p_volumes):
        result *= 0.35
    if q_inches and p_inches and not (q_inches & p_inches):
        result *= 0.45
    return result


def match_query(name, category):
    category = category or ""
    category_norm = norm(category)
    name_norm = norm(name)
    name_tokens = token_set(name)
    if len(name_tokens) <= 4 or name_norm in {"з вимикачем", "без вимикача"}:
        return f"{category} {name}".strip()
    if any(word in category_norm for word in ["фальшдно", "кришк", "пивовар", "ферментер", "флейт"]):
        return f"{category} {name}".strip()
    return str(name)


def clean_url(url):
    parsed = urlparse(urljoin(BASE, url))
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path.rstrip("/"), "", "", ""))


def product_blocks_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    products = []
    for block in soup.select('[data-qaid="product-block"]'):
        title_el = block.select_one(".b-product-gallery__title")
        sku_el = block.select_one(".b-product-gallery__sku")
        if not title_el or not sku_el:
            continue
        title = title_el.get_text(" ", strip=True)
        sku = sku_el.get_text(" ", strip=True)
        href = title_el.get("href")
        if not title or not sku or not href:
            continue
        products.append(Product(title=title, sku=sku, url=clean_url(href)))
    return products, soup


def category_links(soup):
    links = set()
    for anchor in soup.find_all("a", href=True):
        href = clean_url(anchor["href"])
        path = urlparse(href).path
        if not href.startswith(BASE):
            continue
        if re.search(r"/ua/(g|ps)\d+", path) or path.startswith("/ua/product_list"):
            links.add(href)
    return links


def scrape_catalog(max_pages=500):
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
        "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.6",
    })
    products = {}
    seen_pages = set()
    queue = [f"{BASE}/ua/product_list", f"{BASE}/ua"]
    queue.extend(f"{BASE}/ua/product_list/page_{page}" for page in range(2, 12))
    page_index = 0
    while queue and page_index < max_pages:
        url = queue.pop(0)
        url = clean_url(url)
        if url in seen_pages:
            continue
        seen_pages.add(url)
        page_index += 1
        response = session.get(url, timeout=30)
        if response.status_code >= 400:
            continue
        page_products, soup = product_blocks_from_html(response.text)
        before = len(products)
        for product in page_products:
            products[product.url] = product
        for href in category_links(soup):
            if href not in seen_pages:
                queue.append(href)
        print(
            f"scraped page {page_index}: +{len(products) - before} unique, total {len(products)}",
            flush=True,
        )
        time.sleep(0.25)
    return list(products.values())


def load_catalog():
    OUT_DIR.mkdir(exist_ok=True)
    if CATALOG_JSON.exists():
        with CATALOG_JSON.open("r", encoding="utf-8") as fh:
            return [Product(**item) for item in json.load(fh)]
    products = scrape_catalog()
    with CATALOG_JSON.open("w", encoding="utf-8") as fh:
        json.dump([p.__dict__ for p in products], fh, ensure_ascii=False, indent=2)
    return products


def best_match(name, products):
    ranked = sorted(((score(name, p), p) for p in products), key=lambda x: x[0], reverse=True)
    return ranked[0], ranked[1] if len(ranked) > 1 else (None, None)


def status_for(name, best, second):
    best_score, product = best
    second_score = second[0] if second and second[0] is not None else 0
    qn = norm(name)
    pn = norm(product.title)
    if qn == pn:
        return "правильний"
    if qn in pn and best_score >= 0.92 and best_score - second_score >= 0.03:
        return "правильний"
    if best_score >= 0.985 and best_score - second_score >= 0.04:
        return "правильний"
    return "Не точно"


def build_matches(products):
    wb = openpyxl.load_workbook(PRICE_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    current_category = ""
    for row in range(5, ws.max_row + 1):
        category = ws.cell(row, 3).value
        if category:
            current_category = str(category)
        name = ws.cell(row, 4).value
        if not name:
            continue
        query = match_query(name, current_category)
        best, second = best_match(query, products)
        rows.append({
            "row": row,
            "category": current_category,
            "price_name": str(name),
            "query": query,
            "sku": best[1].sku,
            "url": best[1].url,
            "status": status_for(query, best, second),
            "score": round(best[0], 4),
            "matched_title": best[1].title,
            "second_score": round(second[0], 4) if second and second[0] is not None else "",
            "second_title": second[1].title if second and second[1] is not None else "",
        })
    return rows


def write_report(rows):
    OUT_DIR.mkdir(exist_ok=True)
    with MATCH_CSV.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def apply_to_workbook(rows):
    backup = PRICE_FILE.with_name(PRICE_FILE.stem + ".backup_before_sku_url.xlsx")
    if not backup.exists():
        shutil.copy2(PRICE_FILE, backup)
    tmp = PRICE_FILE.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(PRICE_FILE, "r") as zin:
        sheet_xml = zin.read("xl/worksheets/sheet1.xml")
        root = ET.fromstring(sheet_xml)
        ns_uri = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns = {"a": ns_uri}
        ET.register_namespace("", ns_uri)
        sheet_data = root.find("a:sheetData", ns)
        row_nodes = {int(node.attrib["r"]): node for node in sheet_data.findall("a:row", ns)}

        def col_number(cell_ref):
            letters = re.match(r"([A-Z]+)", cell_ref).group(1)
            number = 0
            for letter in letters:
                number = number * 26 + ord(letter) - 64
            return number

        def set_inline(row_number, column, value):
            row_node = row_nodes.get(row_number)
            if row_node is None:
                row_node = ET.SubElement(sheet_data, f"{{{ns_uri}}}row", {"r": str(row_number)})
                row_nodes[row_number] = row_node
            ref = f"{column}{row_number}"
            cell = None
            for candidate in row_node.findall("a:c", ns):
                if candidate.attrib.get("r") == ref:
                    cell = candidate
                    break
            if cell is None:
                cell = ET.Element(f"{{{ns_uri}}}c", {"r": ref})
                existing = list(row_node)
                insert_at = len(existing)
                target_col = col_number(ref)
                for index, candidate in enumerate(existing):
                    if col_number(candidate.attrib.get("r", "XFD1")) > target_col:
                        insert_at = index
                        break
                row_node.insert(insert_at, cell)
            style = cell.attrib.get("s")
            cell.clear()
            cell.attrib["r"] = ref
            if style is not None:
                cell.attrib["s"] = style
            cell.attrib["t"] = "inlineStr"
            is_node = ET.SubElement(cell, f"{{{ns_uri}}}is")
            text_node = ET.SubElement(is_node, f"{{{ns_uri}}}t")
            text_node.text = str(value)

        for item in rows:
            row = item["row"]
            set_inline(row, "I", item["sku"])
            set_inline(row, "J", item["url"])
            set_inline(row, "K", item["status"])

        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
            for info in zin.infolist():
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True) if info.filename == "xl/worksheets/sheet1.xml" else zin.read(info.filename)
                zout.writestr(info, data)
    shutil.move(tmp, PRICE_FILE)
    return backup


def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "report"
    products = load_catalog()
    rows = build_matches(products)
    write_report(rows)
    counts = {"products": len(products), "rows": len(rows)}
    counts["correct"] = sum(1 for row in rows if row["status"] == "правильний")
    counts["uncertain"] = sum(1 for row in rows if row["status"] == "Не точно")
    if mode == "apply":
        counts["backup"] = str(apply_to_workbook(rows))
    print(json.dumps(counts, ensure_ascii=False, indent=2))
    print(f"report={MATCH_CSV}")


if __name__ == "__main__":
    main()
